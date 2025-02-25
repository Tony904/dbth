from PyQt5 import QtCore as qtc
from PyQt5 import QtGui as qtg
from PyQt5 import QtWidgets as qtw
import os
import openai
import base64
import cv2
import numpy as np
from prompts import Prompts
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import platform
import logging
from apikey import mykey
from datetime import datetime

openai.api_key = mykey
DIM = 768 * 2

class DBTHai(qtc.QObject):
    sgl_msg = qtc.pyqtSignal(str)
    sgl_reading_col = qtc.pyqtSignal(str)
    sgl_read_target = qtc.pyqtSignal(list)
    sgl_read_actual = qtc.pyqtSignal(list)
    sgl_read_delta = qtc.pyqtSignal(list)
    sgl_read_lost = qtc.pyqtSignal(list)
    sgl_read_notes = qtc.pyqtSignal(list)
    sgl_saved_to_excel = qtc.pyqtSignal(str)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.excel_path = ''
        self.include_refimg = True
        self.refimg_path = 'C:/Users/jabil/Desktop/dbth/colimgs/_numcol2.png'
        self.pad_left_or_right = 'left'
        self.target_prompt = Prompts.target_prompt_left if self.pad_left_or_right == 'left' else Prompts.target_prompt_right
        self.actual_prompt = Prompts.actual_prompt_left if self.pad_left_or_right == 'left' else Prompts.actual_prompt_right
        self.delta_prompt = Prompts.delta_prompt_left if self.pad_left_or_right == 'left' else Prompts.actual_prompt_right
        self.lost_prompt = Prompts.lost_prompt_left if self.pad_left_or_right == 'left' else Prompts.lost_prompt_right
        self.notes_prompt = Prompts.notes_prompt_left
        self.enabled = True
        self.abort = False
        self.model = Models(Models.GPT_4O)
        self.col_size = 24            

    def send_api_messages(self, colimgs :tuple, machine :str):
        colimgs = self.prepare_imgs(colimgs)
        entarget, enactual, endelta, enlost, ennotes = self.encode_colimgs(colimgs)
        if self.abort: return 1
        self.sgl_reading_col.emit('target')
        target_entries = self.fix_extra_entries(self.send_col_message(entarget, self.target_prompt))
        self.sgl_read_target.emit(target_entries)
        if self.abort: return 1
        self.sgl_reading_col.emit('actual')
        actual_entries = self.fix_extra_entries(self.send_col_message(enactual, self.actual_prompt))
        self.sgl_read_actual.emit(actual_entries)
        if self.abort: return 1
        self.sgl_reading_col.emit('delta')
        delta_entries = self.fix_extra_entries(self.send_col_message(endelta, self.delta_prompt))
        self.sgl_read_delta.emit(delta_entries)
        if self.abort: return 1
        self.sgl_reading_col.emit('lost')
        lost_entries = self.fix_extra_entries(self.send_col_message(enlost, self.lost_prompt))
        self.sgl_read_lost.emit(lost_entries)
        if self.abort: return 1
        self.sgl_reading_col.emit('notes')
        notes_entries = self.fix_extra_entries(self.send_col_message(ennotes, self.notes_prompt))
        self.sgl_read_notes.emit(self.fix_notes(notes_entries))
        if self.abort: return 1
        if machine == '': return 0
        self.export_to_excel(target_entries, actual_entries, delta_entries, lost_entries, notes_entries, machine)
        self.xlog(f'Saved to {self.excel_path}', logging.INFO)
        return 0

    def fix_notes(self, notes):
        if len(notes) < 3: return
        if len(notes) == self.col_size - 1:
            notes.append('')
        if len(notes) == self.col_size:
            if notes[-1] == '' or notes[-1] == 'x':
                if '6-s clean' in notes[self.col_size - 2]:
                    notes[-1] = notes[-2]
                    notes[-2] = 'x'
        return notes

    def fix_extra_entries(self, ent):
        if len(ent) <= self.col_size:
            return ent
        f = list(ent)
        offset = 0
        i = self.col_size - 1
        while i < len(ent):
            s = ent[i]
            if s == 'x': offset += 1
            else: 
                f[i] = f[i - offset]
                f[i - offset] = s
            i += 1
        i = len(f) - 1
        offset = 0
        while i - offset + 1 > self.col_size:
            if f[i - offset] == 'x': offset += 1
            else: break
        return list(f[:i - offset + 1])
    
    def export_to_excel(self, target :list, actual :list, delta :list, lost :list, notes :list, sheet_name :str):
        hours = ['11pm', '12am', '1am', '2am', '3am', '4am', '5am', '6am', '7am', '8am', '9am', '10am', '11am', '12pm', '1pm', '2pm', '3pm', '4pm', '5pm', '6pm', '7pm', '8pm', '9pm', '10pm']
        data = {
            'Hours' : hours,
            'Target': target,
            'Actual': actual,
            'Delta': delta,
            'Lost': lost,
            'Notes': notes
            }
        max_length = 0
        for k in data.keys():
            max_length = len(data[k]) if len(data[k]) > max_length else max_length
        for k in data.keys():
            if len(data[k]) < max_length:
                lst = [''] * (max_length - len(data[k]))
                data[k].extend(lst)
        path = self.excel_path
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.insert_rows(1, amount=max_length + 2)
        else:
            ws = wb.create_sheet(sheet_name)
        headers = list(data.keys())
        for j in range(len(data.keys())):
            ws.cell(row=1, column=j+1, value=headers[j])
            key = headers[j]
            vals = data[key]
            for i in range(0, max_length):
                if key == 'Target' or key == 'Actual' or key == 'Delta':
                    try:
                        ws.cell(row=i+2, column=j+1, value=int(vals[i]))
                    except ValueError:
                        ws.cell(row=i+2, column=j+1, value=vals[i])
                else:
                    ws.cell(row=i+2, column=j+1, value=vals[i])
        ws.cell(row=1, column=7, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        try:
            wb.save(path)
        except PermissionError as e:
            msg = f'{e}\n*** If excel file is open please close and run again. ***'
            self.xlog(msg, logging.ERROR)
            print(msg)
            exit(1)
        # Open excel file
        # if platform.system() == 'Windows':
        #     os.startfile(path)
        # else:
        #     os.system(f'xdg-open "{path}"')

    def send_col_message(self, enimg, prompt):
        messages = [
            {'role': 'system',
            'content': prompt},
            {'role': 'user',
            'content':  [
                {'type': 'image_url',
                'image_url':  {
                    'url': f'data:image/png;base64,{enimg}',
                    'detail': 'high'}}]}]
        response = self.model.send(messages)
        entries = response.choices[0].message.content.split('\n')
        i = 0
        while i < len(entries):
            entries[i] = entries[i].strip()
            if entries[i] == 'X':
                entries[i] = 'x'
            if '```' in entries[i]:
                entries.pop(i)
            else:
                i += 1
        return entries

    def encode_colimgs(self, colimgs :tuple):
        targetimg, actualimg, deltaimg, lostimg, notesimg = colimgs
        dir = os.getcwd() + '/colimgs/'
        targetpath = dir + 'target.png'
        actualpath = dir + 'actual.png'
        deltapath = dir + 'delta.png'
        lostpath = dir + 'lost.png'
        notespath = dir + 'notes.png'
        cv2.imwrite(targetpath, targetimg)
        cv2.imwrite(actualpath, actualimg)
        cv2.imwrite(deltapath, deltaimg)
        cv2.imwrite(lostpath, lostimg)
        cv2.imwrite(notespath, notesimg)
        target64 = self.encode_image(targetpath)
        actual64 = self.encode_image(actualpath)
        delta64 = self.encode_image(deltapath)
        lost64 = self.encode_image(lostpath)
        notes64 = self.encode_image(notespath)
        return target64, actual64, delta64, lost64, notes64

    @staticmethod
    def encode_image(img_path):
        with open(img_path, 'rb') as img_file:
            imgdata = img_file.read()
            return base64.b64encode(imgdata).decode('utf-8')

    def prepare_imgs(self, imgs :tuple) -> list:
        # assume height > width and imgs is in correct order according to dbth table
        prepimgs = []
        for img, lbl in imgs:
            scale = DIM / img.shape[0]
            img = cv2.resize(img, (0, 0), fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
            w = img.shape[1]
            side = self.pad_left_or_right
            if lbl == 'notes':
                side = 'left'
            img, _, _ = self.pad_image_to_square(img, side)
            if self.include_refimg:
                refimg = cv2.imread(self.refimg_path)
                scale = DIM / refimg.shape[0]
                refimg = cv2.resize(refimg, (0, 0), fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
                if side == 'left':
                    img[0:DIM,img.shape[1]-w-refimg.shape[1]:img.shape[1]-w] = refimg.copy()
                else:
                    img[0:DIM,w:w+refimg.shape[1]] = refimg.copy()
            prepimgs.append(img)
        return prepimgs
    
    @staticmethod
    def pad_image_to_square(src, side):
        image = src.copy()
        w = image.shape[1]
        h = image.shape[0]
        top_pad = 0
        bottom_pad = 0
        left_pad = 0
        right_pad = 0
        padded = None
        if w < h:
            if side == 'left':
                left_pad = h - w
            else:
                right_pad = h - w
            padded = np.pad(image, ((top_pad, bottom_pad), (left_pad, right_pad), (0, 0)), mode='constant')
        elif h < w:
            bottom_pad = w - h
            padded = np.pad(image, ((top_pad, bottom_pad), (left_pad, right_pad), (0, 0)), mode='constant')
        else:
            padded = image
        return padded, bottom_pad, right_pad
    
    def xlog(self, msg: str, level :int = logging.DEBUG):
        pass


class Models:
    GPT_4O = 0
    O1 = 1
    O3_MINI = 2

    def __init__(self, model :int = 0):
        self.list = ['gpt-4o', 'o1', 'o3-mini']
        self.selected = model

    def send(self, msg):
        if self.selected == 0:
            response = openai.chat.completions.create(model=self.list[self.selected], messages=msg, temperature=0)
        elif self.selected == 1 or self.selected == 2:
            response = openai.chat.completions.create(model=self.list[self.selected], messages=msg)
        else:
            raise ValueError('Invalid model selected.')
        return response


if __name__ == "__main__":
    models = openai.models.list()
    for model in models:
        print(model.id)